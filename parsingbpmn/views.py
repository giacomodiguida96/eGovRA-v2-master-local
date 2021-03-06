import csv
import json
from datetime import datetime

from django.core.files.storage import FileSystemStorage
from django.core.serializers import python
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from django.http import JsonResponse

from .bpmn_python_master.bpmn_python.bpmn_python_consts import Consts
from .forms import ProcessForm, SystemForm
from .models import Process, Asset, System, Asset_has_attribute, Attribute, Asset_type, Attribute_value, \
    Threat_has_attribute, Threat_has_control, ThreatAgentRiskScores, TACategoryAttribute, ThreatAgentCategory, \
    System_ThreatAgent, TAReplies_Question, TAReplyCategory, Reply, ThreatAgentQuestion, StrideImpactRecord, Stride, \
    Threat_Stride, Risk, OverallRisk, DataObject, Actor
from .bpmn_python_master.bpmn_python import bpmn_diagram_rep as diagram


# Create your views here.

def system_management(request):
    if request.method == 'POST':
        form = SystemForm(request.POST)
        if form.is_valid():
            form.save()
            last_system = System.objects.latest('id')
            return redirect('bpmn_process_management', last_system.pk)
    else:
        form = SystemForm()
    systems = System.objects.all()
    return render(request, 'system_management.html', {
        'form': form, 'systems': systems
    })


def bpmn_process_management(request, systemId):
    pk = systemId
    asset_type = None
    if request.method == 'POST':
        form = ProcessForm(request.POST, request.FILES)
        if form.is_valid():
            saved_form = form.save(commit=False)
            saved_form.system_id = pk
            saved_form.save()
            last_process = Process.objects.latest('id')
            bpmn_graph = diagram.BpmnDiagramGraph()
            pk = last_process.pk
            bpmn_graph.load_diagram_from_xml_file(Process.objects.get(pk=pk).xml)
            #var = Actor.objects.all().get_or_create("",last_process,"")

            lista = bpmn_graph.get_nodes()

            print(lista)
            # print(lista)
            annotations = []
            associations = []
            dataOBJ = []
            dataOutput = []

            for tuple in lista:
                for dizionario in tuple:
                    if type(dizionario) is dict:
                        try:
                            if dizionario['type'].endswith("textAnnotation"):
                                annotations.append(dizionario)
                        except KeyError:
                            print()
                        try:
                            if dizionario['type'].endswith("association"):
                                associations.append(dizionario)
                        except KeyError:
                            print()

                    if type(dizionario) is dict:
                        try:
                            if dizionario['type'].lower().startswith("dataobject"):
                                dataOBJ.append(dizionario["node_name"])
                        except KeyError:
                            print()

            print(dataOBJ)
            print(dataOutput)

            e = ""
            for tuple in lista:
                for dizionario in tuple:
                    if type(dizionario) is dict:
                        if dizionario['type'].lower().endswith("task"):
                            attribute_value = []
                            id_task = dizionario['id']
                            x = dizionario["x"]
                            y = dizionario["y"]
                            width = dizionario["width"]
                            height = dizionario["height"]
                            process_id_bomn = dizionario["process"]

                            #dataOutput = dizionario["dataOutputAssociation"]["id"]
                            #target_refOutput = dizionario["dataOutputAssociation"]["targetRef"]


                            # print(dataOutput)
                            asset_type = None
                            position = x + ":" + y + ":" + width + ":" + height
                            if dizionario['type'].startswith("send"):
                                asset_type = Asset_type.objects.get(name="Send task")
                                e = ""
                                for assoc in associations:
                                    if (id_task == assoc['association'][1]):
                                        target_ref = assoc['association'][2]
                                        for textAnn in annotations:
                                            if (target_ref == textAnn['id']):
                                                e = (textAnn['text'])
                                    e = e.replace(" ", "_")
                                    e = e.lower()
                                if e == "pec_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="PEC communication"))
                                elif e == "mail_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="Mail communication"))
                                elif e == "post_office_communication":
                                    attribute_value.append(
                                        Attribute_value.objects.get(value="Post office communication"))
                            elif dizionario['type'].startswith("receive"):
                                asset_type = Asset_type.objects.get(name="Receive task")
                                id_task = dizionario['id']
                                e = ""
                                for assoc in associations:
                                    if (id_task == assoc['association'][1]):
                                        target_ref = assoc['association'][2]
                                        for textAnn in annotations:
                                            if (target_ref == textAnn['id']):
                                                e = (textAnn['text'])
                                    e = e.replace(" ", "_")
                                    e = e.lower()
                                if e == "pec_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="PEC communication"))
                                elif e == "mail_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="Mail communication"))
                                elif e == "post_office_communication":
                                    attribute_value.append(
                                        Attribute_value.objects.get(value="Post office communication"))
                            elif dizionario['type'].startswith("user"):
                                asset_type = Asset_type.objects.get(name="User task")
                                id_task = dizionario['id']
                                e = ""
                                for assoc in associations:
                                    if (id_task == assoc['association'][1]):
                                        target_ref = assoc['association'][2]
                                        for textAnn in annotations:
                                            print(textAnn)
                                            if (target_ref == textAnn['id']):
                                                e = (textAnn['text'])
                                                print(e)
                                    e = e.replace(" ", "_")
                                    e = e.lower()
                                if e == "online":
                                    attribute_value.append(Attribute_value.objects.get(value="Online"))
                                elif e == "offline":
                                    attribute_value.append(Attribute_value.objects.get(value="Offline"))
                            elif dizionario['type'].startswith("manual"):
                                asset_type = Asset_type.objects.get(name="Manual task")
                                attribute_value.append(Attribute_value.objects.get(value="Manual task"))
                            elif dizionario['type'].startswith("service"):
                                asset_type = Asset_type.objects.get(name="Service task")
                                id_task = dizionario['id']
                                e = ""
                                for assoc in associations:
                                    if (id_task == assoc['association'][1]):
                                        target_ref = assoc['association'][2]
                                        for textAnn in annotations:
                                            print(textAnn)
                                            if (target_ref == textAnn['id']):
                                                e = (textAnn['text'])
                                                print(e)
                                    e = e.replace(" ", "_")
                                e = e.lower()
                                if e == "statefull":
                                    attribute_value.append(Attribute_value.objects.get(value="Statefull"))
                                elif e == "stateless":
                                    attribute_value.append(Attribute_value.objects.get(value="Stateless"))
                            elif dizionario['type'].startswith("script"):
                                asset_type = Asset_type.objects.get(name="Script task")
                                attribute_value.append(Attribute_value.objects.get(value="Script task"))
                            elif dizionario['type'].startswith("business"):
                                asset_type = Asset_type.objects.get(name="Business rule task")
                                attribute_value.append(Attribute_value.objects.get(value="Business rule task"))
                            asset = Asset(name=dizionario['node_name'], bpmn_id=id_task, position=position,
                                          process=Process.objects.get(pk=pk), asset_type=asset_type,process_bpmn_id=process_id_bomn)
                            asset.save()
                            attribute = []
                            for value in attribute_value:
                                attribute.append(Attribute.objects.get(asset_type=asset_type, attribute_value=value))
                            for a in attribute:
                                asset_has_attribute = Asset_has_attribute(asset=asset, attribute=a)
                                asset_has_attribute.save()
                        elif dizionario['type'].endswith("task"):
                            asset = Asset(name=dizionario['node_name'], process=Process.objects.get(pk=pk))
                            asset.save()
                        elif dizionario['type'].lower().startswith("dataobjectreference"):
                            asset_type = Asset_type.objects.get(name="DataObject")
                            try:
                                x = dizionario["x"]
                                y = dizionario["y"]
                                width = dizionario["width"]
                                height = dizionario["height"]
                                process_id_bomn = dizionario["process"]
                                position = x + ":" + y + ":" + width + ":" + height
                            except KeyError:
                                print()
                            asset = Asset(name=dizionario['node_name'], bpmn_id=dizionario["id"], position=position,
                                          process=Process.objects.get(pk=pk), asset_type=asset_type,process_bpmn_id=process_id_bomn)

                            asset.save()


            return redirect('process_data_object_input', systemId, pk)
    else:
        form = ProcessForm()
    processes = Process.objects.filter(system=System.objects.get(pk=pk))
    check_box = []
    for process in processes:
        assets = Asset.objects.filter(process=process)
        check_attribute = False
        for asset in assets:
            if not Asset_has_attribute.objects.filter(asset=asset):
                check_attribute = True
        check_box.append(check_attribute)
    processes_info = zip(processes, check_box)
    return render(request, 'bpmn_process_management.html', {
        'form': form, 'processes_info': processes_info, 'pk': pk, 'systemId': systemId, 'processes': processes
    })


def delete_system(request, systemId):
    if request.method == 'POST':
        system = System.objects.get(id=systemId)
        system.delete()
    return redirect('system_management')


def delete_process(request, systemId, processId):
    if request.method == 'POST':
        process = Process.objects.get(id=processId)
        process.delete()
    return redirect('bpmn_process_management', systemId)


def delete_Dataobj(request, systemId, processID, assetId):
    if request.method == 'POST':
        asset = Asset_type.objects.get(id=assetId)
        asset.delete()
        asset = Asset_type.objects.get(id=assetId + 1)
        asset.delete()
    return redirect('process_data_object_input', systemId)


def process_view_task_type(request, systemId, processId):
    pk = processId
    task_list = Asset.objects.filter(process=Process.objects.get(pk=pk))
    check_attribute = False
    for task in task_list:
        if task.asset_type == None and task.asset_type.lower().endswith("task"):
            check_attribute = True
    if check_attribute == True:
        asset_type = Asset_type.objects.all()
        system = Process.objects.get(id=processId).system
        processes = Process.objects.filter(system=system)
        return render(request, 'process_view_task_type.html', {
            'task_list': task_list, 'asset_type': asset_type, 'systemId': systemId, 'processId': processId,
            'processes': processes
        })
    else:
        return redirect('process_view_attribute', systemId, processId)


def task_type_enrichment(request, systemId, processId):
    if request.method == "POST":
        assets_for_process = Asset.objects.filter(process=Process.objects.get(pk=processId))
        task_enrichment = []
        types = []
        for asset in assets_for_process:
            task_enrichment.append(request.POST.get(str(asset.pk)))
        for type in task_enrichment:
            if type != None:
                type = int(type)
                types.append(Asset_type.objects.get(pk=type))
            else:
                types.append(None)
        for asset, type in zip(assets_for_process, types):
            if type != None:
                x = Asset.objects.get(pk=asset.pk)
                x.asset_type = type
                x.save()
        return redirect('process_view_attribute', systemId, processId)
    else:
        return redirect('task_type_enrichment', systemId, processId)


def process_view_attribute(request, systemId, processId):
    task_list = Asset.objects.filter(process=Process.objects.get(pk=processId)).exclude(asset_type=8)
    # print(task_list)
    check_attribute = False
    for task in task_list:
        if not Asset_has_attribute.objects.filter(asset=task):
            check_attribute = True

    if check_attribute == True:
        task_attributes = []
        list_attributes = []
        for task in task_list:
            task_attributes.append(Asset_has_attribute.objects.filter(asset=task))
        for attributes in task_attributes:
            if not attributes:
                list_attributes.append("empty")
            else:
                sub_list = []
                for element in attributes:
                    sub_list.append(element.attribute.attribute_value.value)
                list_attributes.append(sub_list)
        send = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Send task"))
        receive = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Receive task"))
        user = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="User task"))
        manual = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Manual task"))
        service = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Service task"))
        script = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Script task"))
        business = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Business rule task"))
        task_info = zip(task_list, list_attributes)

        system = Process.objects.get(pk=processId).system
        processes = Process.objects.filter(system=system)

        return render(request, 'process_view_attribute.html', {
            'task_info': task_info, 'send': send, 'receive': receive, 'user': user, 'manual': manual,
            'service': service,
            'script': script, 'business': business, 'systemId': systemId, 'processId': processId,
            'processes': processes})
    else:
        return redirect('threats_and_controls', systemId, processId)


def process_enrichment(request, systemId, processId):
    if request.method == "POST":
        pk = processId
        task_list = Asset.objects.filter(process=Process.objects.get(pk=pk))
        pathfile = Process.objects.filter(id=pk)[0].xml
        bpmn_graph = diagram.BpmnDiagramGraph()
        bpmn_graph.load_diagram_from_xml_file(pathfile)
        check_attribute = False
        for task in task_list:
            if not Asset_has_attribute.objects.filter(asset=task):
                check_attribute = True
        if check_attribute == True:
            assets_for_process = Asset.objects.filter(process=Process.objects.get(pk=pk))
            attributes_enrichment = []
            attributes = []
            for asset in assets_for_process:
                attributes_enrichment.append(request.POST.get(str(asset.pk)))
            for attribute_enrichment in attributes_enrichment:
                if attribute_enrichment != None:
                    attribute_enrichment = int(attribute_enrichment)
                    attributes.append(Attribute.objects.get(pk=attribute_enrichment))
                else:
                    attributes.append(None)

            for asset, attribute in zip(assets_for_process, attributes):
                if attribute != None:
                   print(asset.process_bpmn_id,attribute)
                   bpmn_graph.add_textAnnotation_to_diagram(asset.process_bpmn_id,attribute.attribute_value.value, None)
            bpmn_graph.export_xml_file(r"/Users/giacomodiguida/Documents/", "bpmnfinal.bpmn")
            return redirect('threats_and_controls', systemId, processId)
        else:
            assets_for_process = Asset.objects.filter(process=Process.objects.get(pk=pk))
            attributes_enrichment = []
            attributes = []
            for asset in assets_for_process:
                attributes_enrichment.append(request.POST.get(str(asset.pk)))
            for attribute_enrichment in attributes_enrichment:
                if attribute_enrichment != None:
                    attribute_enrichment = int(attribute_enrichment)
                    attributes.append(Attribute.objects.get(pk=attribute_enrichment))
                else:
                    attributes.append(None)

            for asset, attribute in zip(assets_for_process, attributes):
                if attribute != None:
                    Asset_has_attribute.objects.filter(asset=asset).update(attribute=attribute)
            return redirect('threats_and_controls', systemId, processId)
    else:
        return redirect('process_enrichment', systemId, processId)


import random
import string


def get_random_string(length):
    letters = string.ascii_letters
    result_str = ''.join(random.choice(letters) for i in range(length))
    return result_str


def writeTextAnnotation_bpmn(pathfile, position, taskId, attribute_value):
    textAnnotationId = "TextAnnotation_" + get_random_string(7)
    textAnnotation = "\
    <bpmn:textAnnotation id=\"" + str(textAnnotationId) + "\">\n \
    <bpmn:text>" + str(attribute_value) + "</bpmn:text>\n \
    </bpmn:textAnnotation>\n"
    associationId = "Association_" + get_random_string(7)
    association = "<bpmn:association id=\"" + str(
        associationId) + "\" sourceRef=\"" + taskId + "\" targetRef=\"" + textAnnotationId + "\" />"
    positionValues = position.split(":")
    x = positionValues[0]
    y = positionValues[1]
    width = positionValues[2]
    height = positionValues[3]

    stringToWrite = str(textAnnotation) + " " + str(association)
    shapetextAnn = "<bpmndi:BPMNShape id=\"" + textAnnotationId + "_di\" bpmnElement=" + textAnnotationId + ">\n\
        <dc:Bounds x=\"" + str(int(x) + 20) + "\" y=\"" + str(int(y) + 20) + "\" width=\"" + str(
        width) + "\" height=\"" + height + "\" />\n\
      </bpmndi:BPMNShape>\n"

    shapeAssoc = "<bpmndi:BPMNShape id=\"" + associationId + "_di\" bpmnElement=" + associationId + ">\n\
        <dc:Bounds x=\"" + x + "\" y=\"" + y + "\"/>\n\
        <dc:Bounds x=\"" + str(int(x) + 20) + "\" y=\"" + str(int(y) + 20) + "\"/>\n\
      </bpmndi:BPMNShape>\n"

    f = open(str(pathfile), "r+")
    stringFile = f.read()

    from xml.dom.minidom import parse, parseString

    datasource = open(str(pathfile))  # convert to minidom object
    minidomObject = parse(datasource)
    process = minidomObject.getElementsByTagName('bpmn:process')
    for e in process:
        if (taskId in e.toxml()):
            print("per il task " + str(taskId) + " il process ?? ")
            print(e.toxml())

    # task[0].firstChild.nodeValue

    # Process_1j43nxw

    # print(minidomObject.toxml()) #convert to xml string


def edit_process(request, systemId, processId):
    if request.method == "POST":
        assets = Asset.objects.filter(process=Process.objects.get(pk=processId))
        assets_type = []
        list_attributes = []
        for asset in assets:
            assets_type.append(asset.asset_type)
            list_attributes.append("empty")

        task_info = zip(assets, list_attributes)
        send = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Send task"))
        receive = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Receive task"))
        user = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="User task"))
        manual = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Manual task"))
        service = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Service task"))
        script = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Script task"))
        business = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Business rule task"))
        system = Process.objects.get(pk=processId).system
        processes = Process.objects.filter(system=system)
        return render(request, 'process_view_attribute.html', {
            'task_info': task_info, 'send': send, 'receive': receive, 'user': user, 'manual': manual,
            'service': service, 'script': script, 'business': business, 'systemId': systemId, 'processId': processId,
            'processes': processes})


def threats_and_controls(request, systemId, processId):
    process = Process.objects.get(pk=processId)
    assets = Asset.objects.filter(process=process)
    attributes = []
    threats = []
    controls = []
    for asset in assets:
        attributes.append(Asset_has_attribute.objects.filter(asset=asset))
    for list_attribute in attributes:
        for attribute in list_attribute:
            attribute = attribute.attribute
            threats.append(Threat_has_attribute.objects.filter(attribute=attribute))
    for threats_of_asset in threats:
        sublist_controls = []
        for threat in threats_of_asset:
            threat = threat.threat
            sublist_controls.append(Threat_has_control.objects.filter(threat=threat))
        controls.append(sublist_controls)

    clear_list_threats = []
    for threat_list in threats:
        for threat in threat_list:
            if threat.threat not in clear_list_threats:
                clear_list_threats.append(threat.threat)

    clear_list_controls = []
    for control_of_asset in controls:
        for control_of_threat in control_of_asset:
            for control in control_of_threat:
                if control.control not in clear_list_controls:
                    clear_list_controls.append(control.control)

    system = Process.objects.get(pk=processId).system
    processes = Process.objects.filter(system=system)
    return render(request, 'threats_and_controls.html', {
        'process_name': process.name, 'clear_list_threats': clear_list_threats,
        'clear_list_controls': clear_list_controls, 'systemId': systemId,
        'processId': processId, 'processes': processes
    })


def threat_modeling(systemId, processId):
    assets = Asset.objects.filter(process=Process.objects.get(pk=processId))
    attributes = []
    threats = []
    controls = []
    for asset in assets:
        attributes.append(Asset_has_attribute.objects.filter(asset=asset))
    for list_attribute in attributes:
        for attribute in list_attribute:
            attribute = attribute.attribute
            threats.append(Threat_has_attribute.objects.filter(attribute=attribute))
    for threats_of_asset in threats:
        sublist_controls = []
        for threat in threats_of_asset:
            threat = threat.threat
            sublist_controls.append(Threat_has_control.objects.filter(threat=threat))
        controls.append(sublist_controls)

    controls_per_asset = []
    for asset in threats:
        list_controls = []
        for threat in asset:
            threat = threat.threat
            controls_per_threat = Threat_has_control.objects.filter(threat=threat)
            for control in controls_per_threat:
                control = control.control
                if control not in list_controls:
                    list_controls.append(control)
        controls_per_asset.append(list_controls)

    threat_model_info = zip(assets, attributes, threats, controls, controls_per_asset)
    system = Process.objects.get(pk=processId).system
    processes = Process.objects.filter(system=system)
    context = {
        'threat_model_info': threat_model_info, 'systemId': systemId, 'processId': processId, 'processes': processes
    }
    return context


def threat_modeling_results(systemId, processId):
    assets = Asset.objects.filter(process=Process.objects.get(pk=processId))
    attributes = []
    threats = []
    controls = []
    for asset in assets:
        attributes.append(Asset_has_attribute.objects.filter(asset=asset))
    for list_attribute in attributes:
        for attribute in list_attribute:
            attribute = attribute.attribute
            threats.append(Threat_has_attribute.objects.filter(attribute=attribute))
    for threats_of_asset in threats:
        sublist_controls = []
        for threat in threats_of_asset:
            threat = threat.threat
            sublist_controls.append(Threat_has_control.objects.filter(threat=threat))
        controls.append(sublist_controls)

    controls_per_asset = []
    for asset in threats:
        list_controls = []
        for threat in asset:
            threat = threat.threat
            controls_per_threat = Threat_has_control.objects.filter(threat=threat)
            for control in controls_per_threat:
                control = control.control
                if control not in list_controls:
                    list_controls.append(control)
        controls_per_asset.append(list_controls)

    threat_model_info = zip(assets, threats)
    system = Process.objects.get(pk=processId).system
    processes = Process.objects.filter(system=system)
    context = {
        'threat_model_info': threat_model_info, 'systemId': systemId, 'processId': processId, 'processes': processes
    }
    return context


def threat_modeling_per_asset(systemId, processId, assetId):
    attributes = []
    threats_with_strides = []
    # attributi dell'asset
    attributes.append(Asset_has_attribute.objects.filter(asset=Asset.objects.get(id=assetId)))

    for list_attribute in attributes:
        for attribute in list_attribute:
            attribute = attribute.attribute
            threats = Threat_has_attribute.objects.filter(attribute=attribute)
            for threat in threats:
                strides = Threat_Stride.objects.filter(threat=threat.threat)
                threats_with_strides.append((threat, strides))

    threat_model_info = threats_with_strides
    context = {
        'threat_model_info': threat_model_info, 'systemId': systemId, 'processId': processId, 'assetId': assetId
    }
    return context


def threat_modeling_view(request, systemId, processId):
    context = threat_modeling(systemId, processId)
    result_available = False

    try:
        if (ThreatAgentRiskScores.objects.get(
                system=System.objects.get(id=systemId)) and StrideImpactRecord.objects.filter(
            process=Process.objects.get(id=processId))):
            result_available = True
    except:
        result_available = False
    context['result_available'] = result_available

    return render(request, 'threat_modeling.html', context)


def export_threat_modeling(request, systemId, processId):
    if request.method == "POST":

        # help: https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{name}-report.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
            name=Process.objects.get(pk=processId).name.replace(" ", "_")
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Threat_modeling_REPORT'
        columns = ['Asset name', 'Asset type', 'Asset attributes', 'Threats', 'Policy per asset']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(name="Times New Roman", size=12, bold=True, color='FF0000')
            cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                 right=Side(border_style="thin", color='FF000000'),
                                 top=Side(border_style="thin", color='FF000000'),
                                 bottom=Side(border_style="thin", color='FF000000'), )

        assets = Asset.objects.filter(process=Process.objects.get(pk=processId))
        attributes = []
        threats = []
        for asset in assets:
            attributes.append(Asset_has_attribute.objects.filter(asset=asset))
        for list_attribute in attributes:
            for attribute in list_attribute:
                attribute = attribute.attribute
                threats.append(Threat_has_attribute.objects.filter(attribute=attribute))

        attributes_list = []
        for attribute in attributes:
            attr_sublist = []
            for element in attribute:
                attr_sublist.append(element.attribute.attribute_value.value)
            attributes_list.append(attr_sublist)

        threats_list = []
        for threat in threats:
            threat_sublist = []
            for element in threat:
                threat_sublist.append(element.threat.name)
            threats_list.append(threat_sublist)

        controls_per_asset = []
        for asset in threats:
            list_controls = []
            for threat in asset:
                threat = threat.threat
                controls_per_threat = Threat_has_control.objects.filter(threat=threat)
                for control in controls_per_threat:
                    control = control.control
                    if control not in list_controls:
                        list_controls.append(control)
            controls_per_asset.append(list_controls)

        for asset, attribute, threat, control in zip(assets, attributes_list, threats_list, controls_per_asset):
            row_num += 1

            if not threat:
                threat0 = ''
            else:
                threat0 = str(threat[0])

            # Define the data for each cell in the row
            row = [
                asset.name,
                asset.asset_type.name,
                str(attribute[0]),
                threat0,
                "CIS." + str(control[0].pk) + " - " + str(control[0])
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                     right=Side(border_style="thin", color='FF000000'),
                                     top=Side(border_style="thin", color='FF000000'),
                                     bottom=Side(border_style="thin", color='FF000000'), )

            count_attr = 0
            old_row = row_num
            while count_attr < len(attribute) - 1:
                count_attr += 1
                row_num += 1

                row = [
                    '',
                    '',
                    str(attribute[count_attr]),
                    ''
                ]

                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                    cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                         right=Side(border_style="thin", color='FF000000'),
                                         top=Side(border_style="thin", color='FF000000'),
                                         bottom=Side(border_style="thin", color='FF000000'), )

            count_threats = 0
            count_controls = 0
            row_num = old_row
            while count_threats < len(threat) - 1 or count_controls < len(control) - 1:
                row_num += 1

                if count_threats < len(threat) - 1 and count_controls < len(control) - 1:
                    count_threats += 1
                    count_controls += 1

                    row = [
                        '',
                        '',
                        '',
                        str(threat[count_threats]),
                        "CIS." + str(control[count_controls].pk) + " - " + str(control[count_controls])
                    ]
                elif count_threats < len(threat) - 1 and not count_controls < len(control) - 1:
                    count_threats += 1

                    row = [
                        '',
                        '',
                        '',
                        str(threat[count_threats]),
                        ''
                    ]
                else:
                    count_controls += 1

                    row = [
                        '',
                        '',
                        '',
                        '',
                        "CIS." + str(control[count_controls].pk) + " - " + str(control[count_controls])
                    ]

                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                    cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                         right=Side(border_style="thin", color='FF000000'),
                                         top=Side(border_style="thin", color='FF000000'),
                                         bottom=Side(border_style="thin", color='FF000000'), )
        # Per effettuare il resize delle celle in base a quella pi?? grande
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value

        workbook.save(response)

        return response


def bpmn_viewer(request, pk):
    process = Process.objects.get(pk=pk)
    return render(request, 'bpmn_viewer.html', {
        'process': process
    })


@csrf_exempt
def risk_analysis(request, systemId, processId, assetId):
    stride_impact_list = []
    risk_list = []
    process = Process.objects.get(id=processId)
    # delete all StrideImpactRecord related to process
    # StrideImpactRecord.objects.filter(process=process).delete()

    save = False
    count = 0
    for info, value in (request.POST).items():
        splittedInfo = info.split('_')
        stride = splittedInfo[0]
        impactInfo = splittedInfo[1]
        if (stride == 'spoofing'):
            strideCategory = 'Spoofing'
        if (stride == 'tampering'):
            strideCategory = 'Tampering'
        if (stride == 'repudiation'):
            strideCategory = 'Repudiation'
        if (stride == 'informationdisclosure'):
            strideCategory = 'Information Disclosure'
        if (stride == 'dos'):
            strideCategory = 'Denial Of Services'
        if (stride == 'elevationofprivileges'):
            strideCategory = 'Elevation of privilege'

        if (impactInfo == 'noncompliance'):
            NonComplianceString = 'Non Compliance'
            NonComplianceValue = value
            stride_impact_list.append((strideCategory, NonComplianceString, NonComplianceValue))
        if (impactInfo == 'financialdamage'):
            FinancialDamageValue = value
            FinancialDamageString = 'Financial Damage'
            stride_impact_list.append((strideCategory, FinancialDamageString, FinancialDamageValue))
        if (impactInfo == 'reputationdamage'):
            ReputationDamageValue = value
            ReputationDamageString = 'Reputation Damage'
            stride_impact_list.append((strideCategory, ReputationDamageString, ReputationDamageValue))
        if (impactInfo == 'privacyviolation'):
            PrivacyViolationValue = value
            PrivacyViolationString = 'Privacy Violation'
            stride_impact_list.append((strideCategory, PrivacyViolationString, PrivacyViolationValue))

        count += 1
        if (count == 4):
            save = True
        if (save):
            strideObject = Stride.objects.get(category=strideCategory)
            strideImpactRecord = StrideImpactRecord.objects.all().get_or_create(process=process,
                                                                                stride=strideObject,
                                                                                financialdamage=FinancialDamageValue,
                                                                                reputationdamage=ReputationDamageValue,
                                                                                noncompliance=NonComplianceValue,
                                                                                privacyviolation=PrivacyViolationValue)
            save = False
            count = 0

    system = System.objects.filter(id=systemId).first()
    process = Process.objects.filter(id=processId).first()

    threat_model = (threat_modeling_per_asset(systemId, processId, assetId))
    # print(threat_model['threat_model_info'])
    threats = threat_model['threat_model_info']

    TAscores = ThreatAgentRiskScores.objects.filter(system=system)[0]
    SIRecords = StrideImpactRecord.objects.filter(process=process)

    # print(TAscores)
    # print(SIRecords)

    LossOfConfidentiality = 0
    LossOfIntegrity = 0
    LossOfAvailability = 0
    LossOfAccountability = 0

    SpoofingRiskCount = 1
    TamperingRiskCount = 1
    RepudiationRiskCount = 1
    InformationDisclosureRiskCount = 1
    DoSRiskCount = 1
    EoPRiskCount = 1

    SpoofingRisk = 0
    TamperingRisk = 0
    RepudiationRisk = 0
    InformationDisclosureRisk = 0
    DoSRisk = 0
    EoPRisk = 0

    for threatwithinfo in threats:
        threat = threatwithinfo[0]
        strides = threatwithinfo[1]

        maxFinancial = 0
        maxReputation = 0
        maxnoncompliance = 0
        maxprivacy = 0

        for SIRecord in SIRecords:
            for Threatstride in strides:
                stride = Threatstride.stride.category
                if (SIRecord.stride.category.lower() == stride.lower()):
                    if (maxFinancial < SIRecord.financialdamage):
                        maxFinancial = SIRecord.financialdamage
                    if (maxReputation < SIRecord.reputationdamage):
                        maxReputation = SIRecord.reputationdamage
                    if (maxnoncompliance < SIRecord.noncompliance):
                        maxnoncompliance = SIRecord.noncompliance
                    if (maxprivacy < SIRecord.privacyviolation):
                        maxprivacy = SIRecord.privacyviolation

        LossOfConfidentiality = 5

        LossOfIntegrity = 5

        LossOfAvailability = 5

        LossOfAccountability = 5

        asset = Asset.objects.get(id=assetId)

        threat_strides_obj = Threat_Stride.objects.filter(threat=threat.threat).first()

        """
        threat_strides_list=Threat_Stride.objects.filter(threat=threat.threat)
        threat_strides_obj=[]
        strides_for_threat=[]
        selectedthreat=""
        for threat_stride in threat_strides_list:
            strides_for_threat.append(threat_stride.stride)
            selectedthreat= threat_stride.threat
        threat_strides_obj.append((selectedthreat,strides_for_threat))
        """
        likelihood = 5
        impact = 5

        likelihood = int(
            TAscores.skill + TAscores.motive + TAscores.opportunity + TAscores.size + threat.threat.owasp_ease_of_discovery + threat.threat.owasp_ease_of_exploit + threat.threat.owasp_intrusion_detection +
            threat.threat.owasp_awareness) / 8

        impact = int(
            threat.threat.loss_of_confidentiality + threat.threat.loss_of_integrity + threat.threat.loss_of_availability + threat.threat.loss_of_accountability + maxFinancial + maxReputation + maxnoncompliance + maxprivacy) / 8

        severity = "MEDIUM"

        severity = calculate_severity(likelihood, impact)

        if (threat_strides_obj.stride.category.lower() == "spoofing"):
            SpoofingRiskCount = SpoofingRiskCount + 1
            if (severity == "VERY LOW"):
                SpoofingRisk = SpoofingRisk + 2
            if (severity == "LOW"):
                SpoofingRisk = SpoofingRisk + 4
            if (severity == "MEDIUM"):
                SpoofingRisk = SpoofingRisk + 6
            if (severity == "HIGH"):
                SpoofingRisk = SpoofingRisk + 8
            if (severity == "CRITICAL"):
                SpoofingRisk = SpoofingRisk + 10

        if (threat_strides_obj.stride.category.lower() == "tampering"):
            TamperingRiskCount = TamperingRiskCount + 1
            if (severity == "VERY LOW"):
                TamperingRisk = TamperingRisk + 2
            if (severity == "LOW"):
                TamperingRisk = TamperingRisk + 4
            if (severity == "MEDIUM"):
                TamperingRisk = TamperingRisk + 6
            if (severity == "HIGH"):
                TamperingRisk = TamperingRisk + 8
            if (severity == "CRITICAL"):
                TamperingRisk = TamperingRisk + 10

        if (threat_strides_obj.stride.category.lower() == "repudiation"):
            RepudiationRiskCount = RepudiationRiskCount + 1
            if (severity == "VERY LOW"):
                RepudiationRisk = RepudiationRisk + 2
            if (severity == "LOW"):
                RepudiationRisk = RepudiationRisk + 4
            if (severity == "MEDIUM"):
                RepudiationRisk = RepudiationRisk + 6
            if (severity == "HIGH"):
                RepudiationRisk = RepudiationRisk + 8
            if (severity == "CRITICAL"):
                RepudiationRisk = RepudiationRisk + 10

        if (threat_strides_obj.stride.category.lower() == "information disclosure"):
            InformationDisclosureRiskCount = InformationDisclosureRiskCount + 1
            if (severity == "VERY LOW"):
                InformationDisclosureRisk = InformationDisclosureRisk + 2
            if (severity == "LOW"):
                InformationDisclosureRisk = InformationDisclosureRisk + 4
            if (severity == "MEDIUM"):
                InformationDisclosureRisk = InformationDisclosureRisk + 6
            if (severity == "HIGH"):
                InformationDisclosureRisk = InformationDisclosureRisk + 8
            if (severity == "CRITICAL"):
                InformationDisclosureRisk = InformationDisclosureRisk + 10

        if (threat_strides_obj.stride.category.lower() == "denial of services"):
            DoSRiskCount = DoSRiskCount + 1
            if (severity == "VERY LOW"):
                DoSRisk = DoSRisk + 2
            if (severity == "LOW"):
                DoSRisk = DoSRisk + 4
            if (severity == "MEDIUM"):
                DoSRisk = DoSRisk + 6
            if (severity == "HIGH"):
                DoSRisk = DoSRisk + 8
            if (severity == "CRITICAL"):
                DoSRisk = DoSRisk + 10

        if (threat_strides_obj.stride.category.lower() == "elevation of privilege"):
            EoPRiskCount = EoPRiskCount + 1
            if (severity == "VERY LOW"):
                EoPRisk = EoPRisk + 2
            if (severity == "LOW"):
                EoPRisk = EoPRisk + 4
            if (severity == "MEDIUM"):
                EoPRisk = EoPRisk + 6
            if (severity == "HIGH"):
                EoPRisk = EoPRisk + 8
            if (severity == "CRITICAL"):
                EoPRisk = EoPRisk + 10

        risk = Risk.objects.get_or_create(system=system, process=process, asset=asset, threat_stride=threat_strides_obj,
                                          likelihood=likelihood, impact=impact, severity=severity,
                                          skill=TAscores.skill, motive=TAscores.motive,
                                          opportunity=TAscores.opportunity, size=TAscores.size,
                                          ease_of_discovery=threat.threat.owasp_ease_of_discovery,
                                          ease_of_exploit=threat.threat.owasp_ease_of_exploit,
                                          intrusion_detection=threat.threat.owasp_intrusion_detection,
                                          awareness=threat.threat.owasp_awareness,
                                          loss_of_confidentiality=threat.threat.loss_of_confidentiality,
                                          loss_of_integrity=threat.threat.loss_of_integrity,
                                          loss_of_availability=threat.threat.loss_of_availability,
                                          loss_of_accountability=threat.threat.loss_of_accountability,
                                          financial=maxFinancial, reputation=maxReputation,
                                          non_compliance=maxnoncompliance, privacy=maxprivacy
                                          )
        risk_list.append(risk)

    SpoofingRiskTot = int(SpoofingRisk / SpoofingRiskCount)
    TamperingRiskTot = int(TamperingRisk / TamperingRiskCount)
    RepudiationRiskTot = int(RepudiationRisk / RepudiationRiskCount)
    InformationDisclosureRiskTot = int(InformationDisclosureRisk / InformationDisclosureRiskCount)
    DoSRiskTot = int(DoSRisk / DoSRiskCount)
    EoPRiskTot = int(EoPRisk / EoPRiskCount)

    SpoofingOverallRiskString = calculate_severity_per_stride(SpoofingRiskTot)
    TamperingOverallRiskString = calculate_severity_per_stride(TamperingRiskTot)
    RepOverallRiskString = calculate_severity_per_stride(RepudiationRiskTot)
    InfOverallRiskString = calculate_severity_per_stride(InformationDisclosureRiskTot)
    DoSOverallRiskString = calculate_severity_per_stride(DoSRiskTot)
    EoPOverallRiskString = calculate_severity_per_stride(EoPRiskTot)

    OverallRisk.objects.get_or_create(spoofing=SpoofingOverallRiskString, tampering=TamperingOverallRiskString,
                                      repudiation=RepOverallRiskString, information=InfOverallRiskString,
                                      dos=DoSOverallRiskString, eop=EoPOverallRiskString, system=system,
                                      process=process, asset=asset)

    # risk_list.sort(key=lambda x: (x.likelihood + x.impact) /2, reverse=True)

    return render(request, 'risk_analysis.html', {"assetName": asset, "systemId": systemId, "processId": processId,
                                                  "assetId": assetId, 'risk_list': risk_list,
                                                  'spoofing': SpoofingOverallRiskString,
                                                  'tampering': TamperingOverallRiskString,
                                                  'repudiation': RepOverallRiskString,
                                                  'informationdis': InfOverallRiskString,
                                                  'dos': DoSOverallRiskString, 'eop': EoPOverallRiskString})


@csrf_exempt
def threat_agent_wizard(request, systemId, processId, assetId):
    context = {}
    # Generate question and related replies
    questions = ThreatAgentQuestion.objects.all()
    questions_replies = TAReplies_Question.objects.all()
    questions_replies_list = []
    for question in questions:
        replies = []
        question_replies_dict = {}
        for reply in questions_replies:
            if question == reply.question:
                replies.append(reply.reply.reply)
        question_replies_dict['question'] = question.question
        question_replies_dict['replies'] = replies
        questions_replies_list.append(question_replies_dict)
    context['questions_replies'] = questions_replies_list
    context['systemId'] = systemId
    context['processId'] = processId
    context['assetId'] = assetId

    try:
        if ThreatAgentRiskScores.objects.get(system=System.objects.get(id=systemId)):
            return redirect('StrideImpact', systemId, processId, assetId)
    except:
        pass
    return render(request, 'threat_agent_wizard.html', context)


def calculate_severity(likelihood, impact):
    severity = "MEDIUM"
    if (impact >= 6):
        if (likelihood < 3):
            severity = "MEDIUM"
        elif (likelihood >= 3 and likelihood < 6):
            severity = "HIGH"
        elif (likelihood > 6):
            severity = "CRITICAL"
    elif (impact >= 3 and impact < 6):
        if (likelihood < 3):
            severity = "LOW"
        elif (likelihood >= 3 and likelihood < 6):
            severity = "MEDIUM"
        elif (likelihood >= 6):
            severity = "HIGH"
    elif (impact < 3):
        if (likelihood < 3):
            severity = "VERY LOW"
        elif (likelihood >= 3 and likelihood < 6):
            severity = "LOW"
        elif (likelihood >= 6):
            severity = "MEDIUM"
    return severity


def calculate_severity_per_stride(risk_per_stride):
    severity = "MEDIUM"
    if (risk_per_stride < 3):
        severity = "VERY LOW"
    if (risk_per_stride >= 3 and risk_per_stride < 5):
        severity = "LOW"
    elif (risk_per_stride >= 5 and risk_per_stride < 7):
        severity = "MEDIUM"
    elif (risk_per_stride >= 7 and risk_per_stride < 9):
        severity = "HIGH"
    elif (risk_per_stride >= 9):
        severity = "CRITICAL"
    return severity


@csrf_exempt
def threat_agent_generation(request, systemId, processId, assetId):
    context = {}

    ThreatAgents = []
    ThreatAgentsPerAsset = []
    # for category in ThreatAgentCategory.objects.all():   #inizializzo la lista finale a tutti i TA
    # ThreatAgents.append(category)
    for reply in request.POST:  # per ogni risposta al questionario
        if (reply != 'csrfmiddlewaretoken'):
            ReplyObject = Reply.objects.filter(reply=reply).get()
            tareplycategories = TAReplyCategory.objects.filter(reply=ReplyObject)
            TAList = []
            for replycategory in tareplycategories.all():  # ogni categoria relativa ad una singola risposta
                # print(replycategory.reply.reply + " "+ replycategory.category.category)
                TAList.append(replycategory.category)
                question = TAReplies_Question.objects.filter(reply=ReplyObject)
            ThreatAgentsPerAsset.append((TAList, question))
    numQ3 = 0
    numQ4 = 0
    # conto il numero di risposte date per Q3 e Q4
    for ThreatAgentsList, question in ThreatAgentsPerAsset:  # per ogni risposta
        questionId = question.get().question.Qid
        if (questionId == "Q3"):
            numQ3 += 1
        if (questionId == "Q4"):
            numQ4 += 1

    i = 0
    j = 0
    ThreatAgentsListTemp = []
    for ThreatAgentsList, question in ThreatAgentsPerAsset:  # per ogni risposta
        questionId = question.get().question.Qid
        if (int(questionId) == 1):
            ThreatAgents = ThreatAgentsList
        if (int(questionId) == 2):
            ThreatAgents = intersection(ThreatAgents, ThreatAgentsList)
        if (int(questionId) == 3):
            if (i == 0):
                ThreatAgentsListTemp = ThreatAgentsList
            elif (i < numQ3):
                ThreatAgentsList = union(ThreatAgentsList, ThreatAgentsListTemp)
                ThreatAgentsListTemp = ThreatAgentsList
            if (i == numQ3 - 1):
                ThreatAgents = intersection(ThreatAgents, ThreatAgentsList)
            i = i + 1

        if (int(questionId) == 4):
            if (j == 0):
                ThreatAgentsListTemp = ThreatAgentsList
                j = j + 1
            elif (j == 1):
                ThreatAgentsListTemp = ThreatAgentsList
                j = j + 1
            elif (j < numQ4):
                ThreatAgentsList = union(ThreatAgentsList, ThreatAgentsListTemp)
                ThreatAgentsListTemp = ThreatAgentsList

    ThreatAgents = intersection(ThreatAgents, ThreatAgentsList)
    ThreatAgentsWithInfo = {}
    for ta in ThreatAgents:
        ThreatAgentsWithInfo[ta] = list(TACategoryAttribute.objects.filter(category=ta))
        System_ThreatAgent.objects.get_or_create(
            system=System.objects.get(id=int(systemId)),
            category=ta
        )
    context = {'ThreatAgents': ThreatAgentsWithInfo}
    context['systemId'] = systemId
    context['processId'] = processId
    context['assetId'] = assetId
    return render(request, 'threat_agent_generation.html', context=context)


def intersection(lst1, lst2):
    lst3 = [value for value in lst1 if value in lst2]
    return lst3


def union(lst1, lst2):
    lst3 = list(set(lst1 + lst2))
    return lst3


@csrf_exempt
def calculate_threat_agent_risks(request, systemId, processId, assetId):
    OWASP_Motive_TOT = 0
    OWASP_Size_TOT = 0
    OWASP_Opportunity_TOT = 0
    OWASP_Skill_TOT = 0
    sommapesi = 0

    for category, risk_value in request.POST.items():
        if (category != 'csrfmiddlewaretoken'):
            TACategory = ThreatAgentCategory.objects.get(category=category)
            # per ogni categoria ottieni i Attribute relativi e calcola i 4 parametri owasp con le formule nella tesi.
            TACategoryAttributes = TACategoryAttribute.objects.filter(category=TACategory)
            OWASP_Motive = 0
            OWASP_Size = 0
            OWASP_Opportunity = 0
            OWASP_Skill = 0
            limits = 0
            intent = 0
            access = 0
            resources = 0
            visibility = 0
            skills = 0

            OWASP_Motives = []

            # scorro gli attributi di category
            for TACategoryAttributeVar in TACategoryAttributes:
                if (TACategoryAttributeVar.attribute.attribute == 'Skills'):
                    skills = TACategoryAttributeVar.attribute.score
                if (TACategoryAttributeVar.attribute.attribute == 'Resources'):
                    resources = TACategoryAttributeVar.attribute.score
                if (TACategoryAttributeVar.attribute.attribute == 'Visibility'):
                    visibility = TACategoryAttributeVar.attribute.score
                if (TACategoryAttributeVar.attribute.attribute == 'Limits'):
                    limits = TACategoryAttributeVar.attribute.score
                if (TACategoryAttributeVar.attribute.attribute == 'Intent'):
                    intent = TACategoryAttributeVar.attribute.score
                if (TACategoryAttributeVar.attribute.attribute == 'Access'):
                    access = TACategoryAttributeVar.attribute.score

            if (risk_value == 'L'):
                risk_valueNum = 1
            if (risk_value == 'M'):
                risk_valueNum = 2
            if (risk_value == 'H'):
                risk_valueNum = 3

            sommapesi = sommapesi + risk_valueNum
            OWASP_Motive = ((((intent / 2) + (limits / 4)) / 2) * 10)
            OWASP_Opportunity = ((((access / 2) + (resources / 6) + (visibility / 4)) / 3) * 10)
            OWASP_Size = (resources / 6) * 10
            OWASP_Skill = (skills / 4) * 10

            OWASP_Motive_TOT += (OWASP_Motive * risk_valueNum)
            OWASP_Opportunity_TOT += OWASP_Opportunity * risk_valueNum
            OWASP_Size_TOT += OWASP_Size * risk_valueNum
            OWASP_Skill_TOT += OWASP_Skill * risk_valueNum

    OWASP_Skill_TOT = int(round(OWASP_Skill_TOT / sommapesi))
    OWASP_Motive_TOT = int(round(OWASP_Motive_TOT / sommapesi))
    OWASP_Size_TOT = int(round(OWASP_Size_TOT / sommapesi))
    OWASP_Opportunity_TOT = int(round(OWASP_Opportunity_TOT / sommapesi))

    system = System.objects.get(id=systemId)

    ScoreAlreadyCreated = ThreatAgentRiskScores.objects.filter(system=system)
    if (not ThreatAgentRiskScores.objects.filter(system=system).exists()):
        obj = ThreatAgentRiskScores.objects.get_or_create(
            system=system,
            skill=OWASP_Skill_TOT,
            size=OWASP_Size_TOT,
            motive=OWASP_Motive_TOT,
            opportunity=OWASP_Opportunity_TOT)

    if (StrideImpactRecord.objects.filter(process=Process.objects.get(id=processId))):
        return redirect('risk_analysis', systemId, processId, assetId)

    if (assetId is None):
        return redirect('risk_analysis_results', systemId, processId, assetId)

    return render(request, 'stride_impact_evaluation.html',
                  {"systemId": systemId, 'processId': processId, "assetId": assetId})


def StrideImpact(request, systemId, processId, assetId):
    if (StrideImpactRecord.objects.filter(process=Process.objects.get(id=processId))):
        return redirect('risk_analysis', systemId, processId, assetId)
    return render(request, 'stride_impact_evaluation.html',
                  {"systemId": systemId, 'processId': processId, "assetId": assetId})


def StrideImpact_Result(request, systemId, processId):
    if (StrideImpactRecord.objects.filter(process=Process.objects.get(id=processId))):
        return redirect('risk_analysis_result', systemId, processId)
    return render(request, 'stride_impact_evaluation_result.html',
                  {"systemId": systemId, 'processId': processId})


@csrf_exempt
def risk_analysis_result(request, systemId, processId):
    system = System.objects.filter(id=systemId).first()
    process = Process.objects.filter(id=processId).first()
    assets = Asset.objects.filter(process=process)

    if (not ThreatAgentRiskScores.objects.get(system=System.objects.get(id=systemId))):
        return redirect('threat_agent_wizard', systemId, processId, assets[0].id)

    if (not StrideImpactRecord.objects.filter(process=Process.objects.get(id=processId))):
        return redirect('StrideImpact_Result', systemId, processId)

    SpoofingRisks = []
    TamperingRisks = []
    RepudiationRisks = []
    InformationDisclosureRisks = []
    DoSRisks = []
    EoPRisks = []

    for asset in assets:
        threat_model = (threat_modeling_per_asset(systemId, processId, asset.id))
        # print(threat_model['threat_model_info'])
        threats = threat_model['threat_model_info']

        TAscores = ThreatAgentRiskScores.objects.filter(system=system)[0]
        SIRecords = StrideImpactRecord.objects.filter(process=process)

        # print(TAscores)
        # print(SIRecords)

        LossOfConfidentiality = 0
        LossOfIntegrity = 0
        LossOfAvailability = 0
        LossOfAccountability = 0

        SpoofingRiskCount = 1
        TamperingRiskCount = 1
        RepudiationRiskCount = 1
        InformationDisclosureRiskCount = 1
        DoSRiskCount = 1
        EoPRiskCount = 1

        SpoofingRisk = 0
        TamperingRisk = 0
        RepudiationRisk = 0
        InformationDisclosureRisk = 0
        DoSRisk = 0
        EoPRisk = 0

        for threatwithinfo in threats:
            threat = threatwithinfo[0]
            strides = threatwithinfo[1]

            maxFinancial = 0
            maxReputation = 0
            maxnoncompliance = 0
            maxprivacy = 0

            for SIRecord in SIRecords:
                for Threatstride in strides:
                    stride = Threatstride.stride.category
                    if (SIRecord.stride.category.lower() == stride.lower()):
                        if (maxFinancial < SIRecord.financialdamage):
                            maxFinancial = SIRecord.financialdamage
                        if (maxReputation < SIRecord.reputationdamage):
                            maxReputation = SIRecord.reputationdamage
                        if (maxnoncompliance < SIRecord.noncompliance):
                            maxnoncompliance = SIRecord.noncompliance
                        if (maxprivacy < SIRecord.privacyviolation):
                            maxprivacy = SIRecord.privacyviolation

            LossOfConfidentiality = 5

            LossOfIntegrity = 5

            LossOfAvailability = 5

            LossOfAccountability = 5

            threat_strides_obj = Threat_Stride.objects.filter(threat=threat.threat).first()

            """
            threat_strides_list=Threat_Stride.objects.filter(threat=threat.threat)
            threat_strides_obj=[]
            strides_for_threat=[]
            selectedthreat=""
            for threat_stride in threat_strides_list:
                strides_for_threat.append(threat_stride.stride)
                selectedthreat= threat_stride.threat
            threat_strides_obj.append((selectedthreat,strides_for_threat))
            """
            likelihood = 5
            impact = 5

            likelihood = int(
                TAscores.skill + TAscores.motive + TAscores.opportunity + TAscores.size + threat.threat.owasp_ease_of_discovery + threat.threat.owasp_ease_of_exploit + threat.threat.owasp_intrusion_detection +
                threat.threat.owasp_awareness) / 8

            impact = int(
                threat.threat.loss_of_confidentiality + threat.threat.loss_of_integrity + threat.threat.loss_of_availability + threat.threat.loss_of_accountability + maxFinancial + maxReputation + maxnoncompliance + maxprivacy) / 8

            severity = "MEDIUM"

            severity = calculate_severity(likelihood, impact)

            if (threat_strides_obj.stride.category.lower() == "spoofing"):
                SpoofingRiskCount = SpoofingRiskCount + 1
                if (severity == "VERY LOW"):
                    SpoofingRisk = SpoofingRisk + 2
                if (severity == "LOW"):
                    SpoofingRisk = SpoofingRisk + 4
                if (severity == "MEDIUM"):
                    SpoofingRisk = SpoofingRisk + 6
                if (severity == "HIGH"):
                    SpoofingRisk = SpoofingRisk + 8
                if (severity == "CRITICAL"):
                    SpoofingRisk = SpoofingRisk + 10

            if (threat_strides_obj.stride.category.lower() == "tampering"):
                TamperingRiskCount = TamperingRiskCount + 1
                if (severity == "VERY LOW"):
                    TamperingRisk = TamperingRisk + 2
                if (severity == "LOW"):
                    TamperingRisk = TamperingRisk + 4
                if (severity == "MEDIUM"):
                    TamperingRisk = TamperingRisk + 6
                if (severity == "HIGH"):
                    TamperingRisk = TamperingRisk + 8
                if (severity == "CRITICAL"):
                    TamperingRisk = TamperingRisk + 10

            if (threat_strides_obj.stride.category.lower() == "repudiation"):
                RepudiationRiskCount = RepudiationRiskCount + 1
                if (severity == "VERY LOW"):
                    RepudiationRisk = RepudiationRisk + 2
                if (severity == "LOW"):
                    RepudiationRisk = RepudiationRisk + 4
                if (severity == "MEDIUM"):
                    RepudiationRisk = RepudiationRisk + 6
                if (severity == "HIGH"):
                    RepudiationRisk = RepudiationRisk + 8
                if (severity == "CRITICAL"):
                    RepudiationRisk = RepudiationRisk + 10

            if (threat_strides_obj.stride.category.lower() == "information disclosure"):
                InformationDisclosureRiskCount = InformationDisclosureRiskCount + 1
                if (severity == "VERY LOW"):
                    InformationDisclosureRisk = InformationDisclosureRisk + 2
                if (severity == "LOW"):
                    InformationDisclosureRisk = InformationDisclosureRisk + 4
                if (severity == "MEDIUM"):
                    InformationDisclosureRisk = InformationDisclosureRisk + 6
                if (severity == "HIGH"):
                    InformationDisclosureRisk = InformationDisclosureRisk + 8
                if (severity == "CRITICAL"):
                    InformationDisclosureRisk = InformationDisclosureRisk + 10

            if (threat_strides_obj.stride.category.lower() == "denial of services"):
                DoSRiskCount = DoSRiskCount + 1
                if (severity == "VERY LOW"):
                    DoSRisk = DoSRisk + 2
                if (severity == "LOW"):
                    DoSRisk = DoSRisk + 4
                if (severity == "MEDIUM"):
                    DoSRisk = DoSRisk + 6
                if (severity == "HIGH"):
                    DoSRisk = DoSRisk + 8
                if (severity == "CRITICAL"):
                    DoSRisk = DoSRisk + 10

            if (threat_strides_obj.stride.category.lower() == "elevation of privilege"):
                EoPRiskCount = EoPRiskCount + 1
                if (severity == "VERY LOW"):
                    EoPRisk = EoPRisk + 2
                if (severity == "LOW"):
                    EoPRisk = EoPRisk + 4
                if (severity == "MEDIUM"):
                    EoPRisk = EoPRisk + 6
                if (severity == "HIGH"):
                    EoPRisk = EoPRisk + 8
                if (severity == "CRITICAL"):
                    EoPRisk = EoPRisk + 10

            risk = Risk.objects.get_or_create(system=system, process=process, asset=asset,
                                              threat_stride=threat_strides_obj,
                                              likelihood=likelihood, impact=impact, severity=severity,
                                              skill=TAscores.skill, motive=TAscores.motive,
                                              opportunity=TAscores.opportunity, size=TAscores.size,
                                              ease_of_discovery=threat.threat.owasp_ease_of_discovery,
                                              ease_of_exploit=threat.threat.owasp_ease_of_exploit,
                                              intrusion_detection=threat.threat.owasp_intrusion_detection,
                                              awareness=threat.threat.owasp_awareness,
                                              loss_of_confidentiality=threat.threat.loss_of_confidentiality,
                                              loss_of_integrity=threat.threat.loss_of_integrity,
                                              loss_of_availability=threat.threat.loss_of_availability,
                                              loss_of_accountability=threat.threat.loss_of_accountability,
                                              financial=maxFinancial, reputation=maxReputation,
                                              non_compliance=maxnoncompliance, privacy=maxprivacy
                                              )

        SpoofingRiskTot = int(SpoofingRisk / SpoofingRiskCount)
        SpoofingRisks.append(SpoofingRiskTot)
        TamperingRiskTot = int(TamperingRisk / TamperingRiskCount)
        TamperingRisks.append(TamperingRiskTot)
        RepudiationRiskTot = int(RepudiationRisk / RepudiationRiskCount)
        RepudiationRisks.append(RepudiationRiskTot)
        InformationDisclosureRiskTot = int(InformationDisclosureRisk / InformationDisclosureRiskCount)
        InformationDisclosureRisks.append(InformationDisclosureRiskTot)
        DoSRiskTot = int(DoSRisk / DoSRiskCount)
        DoSRisks.append(DoSRiskTot)
        EoPRiskTot = int(EoPRisk / EoPRiskCount)
        EoPRisks.append(EoPRiskTot)

        SpoofingOverallRiskString = calculate_severity_per_stride(SpoofingRiskTot)
        TamperingOverallRiskString = calculate_severity_per_stride(TamperingRiskTot)
        RepOverallRiskString = calculate_severity_per_stride(RepudiationRiskTot)
        InfOverallRiskString = calculate_severity_per_stride(InformationDisclosureRiskTot)
        DoSOverallRiskString = calculate_severity_per_stride(DoSRiskTot)
        EoPOverallRiskString = calculate_severity_per_stride(EoPRiskTot)

        OverallRisk.objects.get_or_create(spoofing=SpoofingOverallRiskString, tampering=TamperingOverallRiskString,
                                          repudiation=RepOverallRiskString, information=InfOverallRiskString,
                                          dos=DoSOverallRiskString, eop=EoPOverallRiskString, asset=asset,
                                          process=process, system=system)

        maxSpoofing = SpoofingRisks[0]
        maxTampering = TamperingRisks[0]
        maxRep = RepudiationRisks[0]
        maxInf = InformationDisclosureRisks[0]
        maxDoS = DoSRisks[0]
        maxEoP = EoPRisks[0]

        for i in range(len(SpoofingRisks)):
            if (SpoofingRisks[i] >= maxSpoofing):
                maxSpoofing = SpoofingRisks[i]
            if (TamperingRisks[i] >= maxTampering):
                maxTampering = TamperingRisks[i]
            if (RepudiationRisks[i] >= maxRep):
                maxRep = RepudiationRisks[i]
            if (InformationDisclosureRisks[i] >= maxInf):
                maxInf = InformationDisclosureRisks[i]
            if (DoSRisks[i] >= maxDoS):
                maxDoS = DoSRisks[i]
            if (EoPRisks[i] >= maxEoP):
                maxEoP = EoPRisks[i]

    SpoofingOverallRiskString = calculate_severity_per_stride(maxSpoofing)
    TamperingOverallRiskString = calculate_severity_per_stride(maxTampering)
    RepOverallRiskString = calculate_severity_per_stride(maxRep)
    InfOverallRiskString = calculate_severity_per_stride(maxInf)
    DoSOverallRiskString = calculate_severity_per_stride(maxDoS)
    EoPOverallRiskString = calculate_severity_per_stride(maxEoP)

    return render(request, 'risk_analysis_result.html',
                  {"processName": process, "systemId": systemId, "processId": processId,
                   'spoofing': SpoofingOverallRiskString,
                   'tampering': TamperingOverallRiskString,
                   'repudiation': RepOverallRiskString,
                   'informationdis': InfOverallRiskString,
                   'dos': DoSOverallRiskString, 'eop': EoPOverallRiskString})


def process_data_object_input(request, systemId, processId):
    print(request.POST.get)
    nameprocess = Process.objects.get(pk=processId)
    asset_type = Asset_type.objects.get(name="DataObject")

    lista = Asset.objects.filter(process=Process.objects.get(pk=processId), asset_type=asset_type).exclude(name="")
    list_data = []
    for tuple in lista:
        list_data.append(tuple)
    return render(request, 'process_dataobject_input.html',
                  {"systemId": systemId, "processId": processId, "list_data": list_data})

    for tuple in lista:
        list_data.append(tuple)

    return render(request, 'process_dataobject_input.html',
                  {"systemId": systemId, "processId": processId, "list_data": list_data})






def save_dataobject(request, systemId, processId):
    nameprocess = Process.objects.get(pk=processId)
    asset_type = Asset_type.objects.get(name="DataObject")

    last_process = Process.objects.latest('id')
    bpmn_graph = diagram.BpmnDiagramGraph()
    pk = last_process.pk
    bpmn_graph.load_diagram_from_xml_file(Process.objects.get(pk=pk).xml)
    lista2 = bpmn_graph.get_nodes()
    lista = Asset.objects.filter(process=Process.objects.get(pk=processId), asset_type=asset_type).exclude(name="")
    list_data = []
    task_with_data = []
    post_data = dict(request.POST.lists())
    temp = []

    for tmp in post_data:
        temp.append(tmp)
    #print(temp)
    #print(len(lista), "json")
    for dato in range(len(lista),len(post_data)-1):
        dataref = "DataObjectReference_" + get_random_string(7)
        # dataobj = "DataObject_" + get_random_string(7)
        asset_type = Asset_type.objects.get(name="DataObject")

        asset = Asset(name=temp[dato], bpmn_id=dataref, position="",
                      process=Process.objects.get(pk=processId), asset_type=asset_type)
        asset.save()
    for tuple in lista2:
        for dizionario in tuple:
            if type(dizionario) is dict:
                if dizionario['type'].lower().endswith("task"):
                    try:
                        if dizionario['dataOutputAssociation']["id"] != "":
                            key = dizionario['dataOutputAssociation']["targetRef"]
                            # data_name=Asset.objects.filter(bpmn_id=key)
                            for tuple2 in lista2:
                                for dizionario2 in tuple2:
                                    if type(dizionario2) is dict:
                                        # print(dizionario2["id"], dizionario["targetRef"]["targetRef"])
                                        if dizionario2['id'] == key:
                                            keydata = dizionario2["node_name"]
                                            print(keydata)
                                            task_with_data.append(
                                                {"task": dizionario["node_name"], "data": dizionario2["node_name"]})

                    except KeyError:
                        print()



    print(task_with_data, "lista2")

    # print(request.POST.items, "sto qua")
    print(lista2)
    list_data = Asset.objects.filter(process=Process.objects.get(pk=processId), asset_type=asset_type).exclude(name="")
    list_task = Asset.objects.filter(process=Process.objects.get(pk=processId)).exclude(asset_type=9)
    print(list_task)
    return render(request, 'process_dataobject_enrichment.html',
                  {"systemId": systemId, "processId": processId, "list_data": list_data,"list_task":list_task,"task_with_data":task_with_data})
